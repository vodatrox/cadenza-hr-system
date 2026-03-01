"""
Export utilities for payroll data.
Clean, modern, professional document design.
"""
from io import BytesIO
from datetime import datetime
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
)
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

from .models import PayrollBatch, Payroll


# ── Shared palette ────────────────────────────────────────────────────────────
NAVY = '0F172A'
SLATE_700 = '334155'
SLATE_500 = '64748B'
SLATE_300 = 'CBD5E1'
SLATE_100 = 'F1F5F9'
SLATE_50 = 'F8FAFC'
WHITE = 'FFFFFF'
ACCENT = '0D9488'  # teal
EMERALD = '059669'
RED = 'DC2626'

rc = lambda hex_str: colors.HexColor(f'#{hex_str}')

THIN_BORDER_SIDE = Side(style='thin', color=SLATE_300)
HAIR_BORDER_SIDE = Side(style='hair', color=SLATE_300)
NO_BORDER = Side(style=None)


class PayrollBatchExcelExport:
    """Generate a clean, modern Excel export for payroll batch."""

    def __init__(self, batch):
        self.batch = batch
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = 'Payroll Report'
        self.ws.sheet_properties.tabColor = NAVY

    def generate(self):
        self._setup_page()
        self._add_header()
        self._add_batch_info()
        self._add_summary()
        self._add_payroll_table()
        self._format_columns()

        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output

    def _setup_page(self):
        self.ws.sheet_view.showGridLines = False
        self.ws.page_margins.left = 0.5
        self.ws.page_margins.right = 0.5
        self.ws.page_margins.top = 0.4
        self.ws.page_margins.bottom = 0.4

    def _add_header(self):
        # Company name — large, dark
        cell = self.ws['B2']
        cell.value = self.batch.client.name
        cell.font = Font(name='Calibri', size=18, bold=True, color=NAVY)
        self.ws.merge_cells('B2:E2')

        # Subtitle
        cell = self.ws['B3']
        cell.value = 'Payroll Batch Report'
        cell.font = Font(name='Calibri', size=11, color=SLATE_500)
        self.ws.merge_cells('B3:E3')

        # Date — right aligned
        cell = self.ws['G2']
        cell.value = datetime.now().strftime('%B %d, %Y')
        cell.font = Font(name='Calibri', size=10, color=SLATE_500)
        cell.alignment = Alignment(horizontal='right')
        self.ws.merge_cells('G2:H2')

        # Thin rule under header
        for col in range(2, 9):
            self.ws.cell(row=4, column=col).border = Border(bottom=THIN_BORDER_SIDE)

    def _add_batch_info(self):
        row = 6
        fields = [
            ('Batch', self.batch.batch_number),
            ('Title', self.batch.title),
            ('Period', self.batch.period.title),
            ('Status', self.batch.status),
            ('Employees', str(self.batch.payroll_count)),
        ]

        label_font = Font(name='Calibri', size=9, color=SLATE_500)
        value_font = Font(name='Calibri', size=10, bold=True, color=NAVY)

        for i, (label, value) in enumerate(fields):
            col_label = 2 + (i % 3) * 2
            r = row + (i // 3) * 2
            lc = self.ws.cell(row=r, column=col_label)
            lc.value = label.upper()
            lc.font = label_font
            vc = self.ws.cell(row=r + 1, column=col_label)
            vc.value = value
            vc.font = value_font

    def _add_summary(self):
        row = 12

        # Section label
        cell = self.ws.cell(row=row, column=2)
        cell.value = 'FINANCIAL SUMMARY'
        cell.font = Font(name='Calibri', size=9, bold=True, color=SLATE_500, )
        self.ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        row += 1

        summaries = [
            ('Gross Pay', self.batch.total_gross_pay, NAVY),
            ('Deductions', self.batch.total_deductions, RED),
            ('Net Pay', self.batch.total_net_pay, EMERALD),
        ]

        for i, (label, value, clr) in enumerate(summaries):
            col = 2 + i * 2
            lc = self.ws.cell(row=row, column=col)
            lc.value = label
            lc.font = Font(name='Calibri', size=9, color=SLATE_500)

            vc = self.ws.cell(row=row + 1, column=col)
            vc.value = float(value)
            vc.font = Font(name='Calibri', size=14, bold=True, color=clr)
            vc.number_format = '₦#,##0.00'

        # Divider
        row += 3
        for col in range(2, 9):
            self.ws.cell(row=row, column=col).border = Border(bottom=THIN_BORDER_SIDE)

        self._table_start_row = row + 2

    def _add_payroll_table(self):
        row = self._table_start_row

        headers = ['#', 'Employee ID', 'Employee Name', 'Basic Salary',
                   'Gross Pay', 'Deductions', 'Net Pay']

        header_font = Font(name='Calibri', size=9, bold=True, color=SLATE_500)
        header_fill = PatternFill(start_color=SLATE_50, end_color=SLATE_50, fill_type='solid')

        for col_num, header in enumerate(headers, 2):
            cell = self.ws.cell(row=row, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = Border(bottom=THIN_BORDER_SIDE)
            cell.alignment = Alignment(
                horizontal='right' if col_num >= 5 else 'left',
                vertical='center'
            )

        row += 1

        payrolls = list(
            self.batch.payrolls.all()
            .select_related('employee')
            .order_by('employee__first_name')
        )

        data_font = Font(name='Calibri', size=10, color=NAVY)
        data_font_muted = Font(name='Calibri', size=10, color=SLATE_700)
        alt_fill = PatternFill(start_color=SLATE_50, end_color=SLATE_50, fill_type='solid')
        currency_fmt = '₦#,##0.00'

        for idx, payroll in enumerate(payrolls, 1):
            values = [
                idx,
                payroll.employee.employee_id,
                payroll.employee.get_full_name(),
                float(payroll.basic_salary),
                float(payroll.gross_pay),
                float(payroll.total_deductions),
                float(payroll.net_pay),
            ]

            for col_num, value in enumerate(values, 2):
                cell = self.ws.cell(row=row, column=col_num)
                cell.value = value
                cell.font = data_font if col_num <= 4 else data_font_muted
                cell.border = Border(bottom=Border(bottom=HAIR_BORDER_SIDE).bottom)

                if col_num >= 5:
                    cell.number_format = currency_fmt
                    cell.alignment = Alignment(horizontal='right')

                if idx % 2 == 0:
                    cell.fill = alt_fill

            row += 1

        # Totals row
        row += 1
        total_fill = PatternFill(start_color=SLATE_100, end_color=SLATE_100, fill_type='solid')
        total_font = Font(name='Calibri', size=10, bold=True, color=NAVY)

        self.ws.cell(row=row, column=2).value = ''
        self.ws.cell(row=row, column=3).value = ''
        label_cell = self.ws.cell(row=row, column=4)
        label_cell.value = 'TOTAL'
        label_cell.font = Font(name='Calibri', size=9, bold=True, color=SLATE_500)
        label_cell.alignment = Alignment(horizontal='right')

        totals = [
            float(sum(p.basic_salary for p in payrolls)),
            float(sum(p.gross_pay for p in payrolls)),
            float(sum(p.total_deductions for p in payrolls)),
            float(sum(p.net_pay for p in payrolls)),
        ]

        for i, total in enumerate(totals):
            cell = self.ws.cell(row=row, column=5 + i)
            cell.value = total
            cell.font = total_font
            cell.number_format = currency_fmt
            cell.alignment = Alignment(horizontal='right')
            cell.fill = total_fill
            cell.border = Border(top=THIN_BORDER_SIDE)

    def _format_columns(self):
        widths = {'A': 2, 'B': 5, 'C': 14, 'D': 28, 'E': 16, 'F': 16, 'G': 16, 'H': 16}
        for col, width in widths.items():
            self.ws.column_dimensions[col].width = width

        # Row heights
        self.ws.row_dimensions[2].height = 26
        self.ws.row_dimensions[3].height = 18


class PayrollBatchPDFExport:
    """Generate a clean, modern PDF export for payroll batch."""

    def __init__(self, batch):
        self.batch = batch
        self.styles = getSampleStyleSheet()
        self._setup_styles()

    def _setup_styles(self):
        self.styles.add(ParagraphStyle(
            name='CompanyName',
            fontName='Helvetica-Bold',
            fontSize=18,
            textColor=rc(NAVY),
            spaceAfter=2,
        ))
        self.styles.add(ParagraphStyle(
            name='DocSubtitle',
            fontName='Helvetica',
            fontSize=10,
            textColor=rc(SLATE_500),
            spaceAfter=16,
        ))
        self.styles.add(ParagraphStyle(
            name='SectionLabel',
            fontName='Helvetica-Bold',
            fontSize=8,
            textColor=rc(SLATE_500),
            spaceBefore=20,
            spaceAfter=8,
        ))
        self.styles.add(ParagraphStyle(
            name='MetaLabel',
            fontName='Helvetica',
            fontSize=8,
            textColor=rc(SLATE_500),
        ))
        self.styles.add(ParagraphStyle(
            name='MetaValue',
            fontName='Helvetica-Bold',
            fontSize=10,
            textColor=rc(NAVY),
        ))
        self.styles.add(ParagraphStyle(
            name='FooterText',
            fontName='Helvetica',
            fontSize=7,
            textColor=rc(SLATE_500),
            alignment=TA_CENTER,
        ))

    def generate(self):
        output = BytesIO()
        doc = SimpleDocTemplate(
            output, pagesize=A4,
            rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40
        )

        story = []
        pw = doc.width  # usable page width

        # ── Header ──────────────────────────────────────────────────
        story.append(Paragraph(self.batch.client.name, self.styles['CompanyName']))
        story.append(Paragraph(
            f"Payroll Batch Report &bull; {datetime.now().strftime('%B %d, %Y')}",
            self.styles['DocSubtitle']
        ))
        story.append(HRFlowable(
            width='100%', thickness=0.5, color=rc(SLATE_300),
            spaceAfter=16, spaceBefore=0
        ))

        # ── Batch info ──────────────────────────────────────────────
        meta_items = [
            ('BATCH', self.batch.batch_number),
            ('TITLE', self.batch.title),
            ('PERIOD', self.batch.period.title),
            ('STATUS', self.batch.status),
            ('EMPLOYEES', str(self.batch.payroll_count)),
        ]

        meta_row_1 = []
        meta_row_2 = []
        for label, value in meta_items[:3]:
            meta_row_1.append(Paragraph(f"<font size=7 color='#{SLATE_500}'>{label}</font><br/>"
                                        f"<font size=10><b>{value}</b></font>",
                                        self.styles['Normal']))
        for label, value in meta_items[3:]:
            meta_row_2.append(Paragraph(f"<font size=7 color='#{SLATE_500}'>{label}</font><br/>"
                                        f"<font size=10><b>{value}</b></font>",
                                        self.styles['Normal']))
        # pad row 2 to 3 cols
        while len(meta_row_2) < 3:
            meta_row_2.append('')

        col_w = pw / 3
        meta_table = Table([meta_row_1, meta_row_2], colWidths=[col_w] * 3)
        meta_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ]))
        story.append(meta_table)

        # ── Financial Summary ───────────────────────────────────────
        story.append(Paragraph('FINANCIAL SUMMARY', self.styles['SectionLabel']))

        summary_items = [
            ('Gross Pay', f"₦{self.batch.total_gross_pay:,.2f}", NAVY),
            ('Deductions', f"₦{self.batch.total_deductions:,.2f}", RED),
            ('Net Pay', f"₦{self.batch.total_net_pay:,.2f}", EMERALD),
        ]

        sum_cells = []
        for label, value, clr in summary_items:
            sum_cells.append(Paragraph(
                f"<font size=7 color='#{SLATE_500}'>{label}</font><br/>"
                f"<font size=14 color='#{clr}'><b>{value}</b></font>",
                self.styles['Normal']
            ))

        sum_table = Table([sum_cells], colWidths=[col_w] * 3)
        sum_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), rc(SLATE_50)),
            ('ROUNDEDCORNERS', [4, 4, 4, 4]),
            ('TOPPADDING', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 14),
            ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ]))
        story.append(sum_table)

        # ── Payroll Table ───────────────────────────────────────────
        story.append(Paragraph('PAYROLL DETAILS', self.styles['SectionLabel']))

        payrolls = list(
            self.batch.payrolls.all()
            .select_related('employee')
            .order_by('employee__first_name')
        )

        # Header row
        table_data = [['#', 'Employee', 'Basic Salary', 'Gross Pay', 'Deductions', 'Net Pay']]

        for idx, p in enumerate(payrolls, 1):
            table_data.append([
                str(idx),
                p.employee.get_full_name()[:30],
                f"₦{p.basic_salary:,.2f}",
                f"₦{p.gross_pay:,.2f}",
                f"₦{p.total_deductions:,.2f}",
                f"₦{p.net_pay:,.2f}",
            ])

        # Total row
        table_data.append([
            '', 'Total',
            f"₦{sum(p.basic_salary for p in payrolls):,.2f}",
            f"₦{sum(p.gross_pay for p in payrolls):,.2f}",
            f"₦{sum(p.total_deductions for p in payrolls):,.2f}",
            f"₦{sum(p.net_pay for p in payrolls):,.2f}",
        ])

        col_widths = [0.4 * inch, 2.2 * inch, 1.1 * inch, 1.1 * inch, 1.1 * inch, 1.1 * inch]
        ptable = Table(table_data, colWidths=col_widths, repeatRows=1)

        style_cmds = [
            # Header
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('TEXTCOLOR', (0, 0), (-1, 0), rc(SLATE_500)),
            ('BACKGROUND', (0, 0), (-1, 0), rc(SLATE_50)),
            ('LINEBELOW', (0, 0), (-1, 0), 0.5, rc(SLATE_300)),
            # Data rows
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 9),
            ('TEXTCOLOR', (0, 1), (-1, -2), rc(NAVY)),
            # Alignment
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
            # Padding
            ('TOPPADDING', (0, 0), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            # Row lines
            ('LINEBELOW', (0, 1), (-1, -2), 0.25, rc(SLATE_300)),
            # Total row
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 9),
            ('TEXTCOLOR', (0, -1), (-1, -1), rc(NAVY)),
            ('LINEABOVE', (0, -1), (-1, -1), 1, rc(SLATE_700)),
            ('BACKGROUND', (0, -1), (-1, -1), rc(SLATE_50)),
        ]

        # Alternating row tint
        for i in range(2, len(table_data) - 1, 2):
            style_cmds.append(('BACKGROUND', (0, i), (-1, i), rc(SLATE_50)))

        ptable.setStyle(TableStyle(style_cmds))
        story.append(ptable)

        # Footer
        story.append(Spacer(1, 0.4 * inch))
        story.append(HRFlowable(
            width='100%', thickness=0.25, color=rc(SLATE_300),
            spaceAfter=8, spaceBefore=0
        ))
        story.append(Paragraph(
            f"{self.batch.client.name} &bull; Confidential",
            self.styles['FooterText']
        ))

        doc.build(story)
        output.seek(0)
        return output


class PayslipPDFExport:
    """Generate a sleek, modern PDF payslip."""

    def __init__(self, payroll):
        self.payroll = payroll
        self.styles = getSampleStyleSheet()
        self._setup_styles()

    def _setup_styles(self):
        self.styles.add(ParagraphStyle(
            name='CompanyName',
            fontName='Helvetica-Bold',
            fontSize=16,
            textColor=rc(NAVY),
            spaceAfter=2,
        ))
        self.styles.add(ParagraphStyle(
            name='DocLabel',
            fontName='Helvetica',
            fontSize=8,
            textColor=rc(SLATE_500),
            alignment=TA_RIGHT,
        ))
        self.styles.add(ParagraphStyle(
            name='SectionLabel',
            fontName='Helvetica-Bold',
            fontSize=8,
            textColor=rc(SLATE_500),
            spaceBefore=16,
            spaceAfter=8,
        ))
        self.styles.add(ParagraphStyle(
            name='FooterText',
            fontName='Helvetica',
            fontSize=7,
            textColor=rc(SLATE_500),
            alignment=TA_CENTER,
        ))

    def generate(self):
        output = BytesIO()
        doc = SimpleDocTemplate(
            output, pagesize=letter,
            rightMargin=50, leftMargin=50, topMargin=44, bottomMargin=40
        )

        story = []
        pw = doc.width

        # ── Header ──────────────────────────────────────────────────
        header_data = [[
            Paragraph(self.payroll.client.name, self.styles['CompanyName']),
            Paragraph(
                f"<font size=7 color='#{SLATE_500}'>PAYSLIP</font><br/>"
                f"<font size=10><b>{self.payroll.period.title}</b></font>",
                ParagraphStyle('tmp', parent=self.styles['Normal'], alignment=TA_RIGHT)
            ),
        ]]
        ht = Table(header_data, colWidths=[pw * 0.6, pw * 0.4])
        ht.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))
        story.append(ht)
        story.append(Spacer(1, 4))
        story.append(HRFlowable(
            width='100%', thickness=0.5, color=rc(SLATE_300),
            spaceAfter=16, spaceBefore=4
        ))

        # ── Employee info strip ─────────────────────────────────────
        emp = self.payroll.employee
        payment_date_str = (
            self.payroll.payment_date.strftime('%B %d, %Y')
            if self.payroll.payment_date else 'Pending'
        )

        info_cells = [
            Paragraph(f"<font size=7 color='#{SLATE_500}'>EMPLOYEE</font><br/>"
                      f"<font size=10><b>{emp.get_full_name()}</b></font>",
                      self.styles['Normal']),
            Paragraph(f"<font size=7 color='#{SLATE_500}'>EMPLOYEE ID</font><br/>"
                      f"<font size=10><b>{emp.employee_id}</b></font>",
                      self.styles['Normal']),
            Paragraph(f"<font size=7 color='#{SLATE_500}'>PAYMENT DATE</font><br/>"
                      f"<font size=10><b>{payment_date_str}</b></font>",
                      self.styles['Normal']),
        ]

        info_table = Table([info_cells], colWidths=[pw / 3] * 3)
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), rc(SLATE_50)),
            ('ROUNDEDCORNERS', [4, 4, 4, 4]),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ]))
        story.append(info_table)

        # ── Collect earnings & deductions ───────────────────────────
        earnings = [('Basic Salary', self.payroll.basic_salary)]

        if self.payroll.total_allowances > 0:
            earnings.append(('Allowances', self.payroll.total_allowances))

        from .serializers import PayrollDetailSerializer
        serializer = PayrollDetailSerializer(self.payroll)
        if serializer.data.get('statutory_earnings_breakdown'):
            for e in serializer.data['statutory_earnings_breakdown']:
                earnings.append((e['name'], Decimal(e['amount'])))

        if self.payroll.earnings.exists():
            for e in self.payroll.earnings.all():
                earnings.append((e.name, e.amount))

        deductions = []
        if self.payroll.pension > 0:
            deductions.append(('Pension', self.payroll.pension))
        if self.payroll.nhf > 0:
            deductions.append(('NHF', self.payroll.nhf))
        if self.payroll.tax > 0:
            deductions.append(('Income Tax (PAYE)', self.payroll.tax))

        if serializer.data.get('statutory_deductions_breakdown'):
            for d in serializer.data['statutory_deductions_breakdown']:
                deductions.append((d['name'], Decimal(d['amount'])))

        if self.payroll.deductions.exists():
            for d in self.payroll.deductions.all():
                deductions.append((d.name, d.amount))

        # ── Side-by-side earnings & deductions table ────────────────
        half_w = pw / 2

        # Build a unified table: [earn_label, earn_amt, deduct_label, deduct_amt]
        story.append(Spacer(1, 6))

        # Section headers row
        table_data = [[
            Paragraph(f"<font color='#{EMERALD}'>&bull;</font>&nbsp;&nbsp;"
                      f"<font size=8><b>EARNINGS</b></font>", self.styles['Normal']),
            '',
            Paragraph(f"<font color='#{RED}'>&bull;</font>&nbsp;&nbsp;"
                      f"<font size=8><b>DEDUCTIONS</b></font>", self.styles['Normal']),
            '',
        ]]

        max_rows = max(len(earnings), len(deductions))
        for i in range(max_rows):
            row = []
            if i < len(earnings):
                row.extend([earnings[i][0], f"₦{earnings[i][1]:,.2f}"])
            else:
                row.extend(['', ''])
            if i < len(deductions):
                row.extend([deductions[i][0], f"₦{deductions[i][1]:,.2f}"])
            else:
                row.extend(['', ''])
            table_data.append(row)

        # Totals row
        table_data.append([
            Paragraph("<b>Gross Pay</b>", self.styles['Normal']),
            Paragraph(f"<b>₦{self.payroll.gross_pay:,.2f}</b>", ParagraphStyle(
                'rtmp', parent=self.styles['Normal'], alignment=TA_RIGHT)),
            Paragraph("<b>Total Deductions</b>", self.styles['Normal']),
            Paragraph(f"<b>₦{self.payroll.total_deductions:,.2f}</b>", ParagraphStyle(
                'rtmp2', parent=self.styles['Normal'], alignment=TA_RIGHT)),
        ])

        cw = [half_w * 0.62, half_w * 0.38, half_w * 0.62, half_w * 0.38]
        ed_table = Table(table_data, colWidths=cw)

        ed_styles = [
            # Header row
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('TEXTCOLOR', (0, 0), (-1, 0), rc(SLATE_500)),
            ('LINEBELOW', (0, 0), (1, 0), 0.5, rc(SLATE_300)),
            ('LINEBELOW', (2, 0), (3, 0), 0.5, rc(SLATE_300)),
            # Data styling
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 9),
            ('TEXTCOLOR', (0, 1), (0, -2), rc(SLATE_700)),
            ('TEXTCOLOR', (2, 1), (2, -2), rc(SLATE_700)),
            ('TEXTCOLOR', (1, 1), (1, -2), rc(NAVY)),
            ('TEXTCOLOR', (3, 1), (3, -2), rc(NAVY)),
            # Right-align amounts
            ('ALIGN', (1, 1), (1, -2), 'RIGHT'),
            ('ALIGN', (3, 1), (3, -2), 'RIGHT'),
            # Padding
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            # Light row dividers
            ('LINEBELOW', (0, 1), (1, -2), 0.25, rc(SLATE_300)),
            ('LINEBELOW', (2, 1), (3, -2), 0.25, rc(SLATE_300)),
            # Gap between columns
            ('LEFTPADDING', (2, 0), (2, -1), 16),
            # Totals row
            ('LINEABOVE', (0, -1), (-1, -1), 1, rc(SLATE_700)),
            ('BACKGROUND', (0, -1), (-1, -1), rc(SLATE_50)),
            ('TOPPADDING', (0, -1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, -1), (-1, -1), 10),
        ]

        ed_table.setStyle(TableStyle(ed_styles))
        story.append(ed_table)
        story.append(Spacer(1, 16))

        # ── Net Pay card ────────────────────────────────────────────
        net_pay_data = [[
            Paragraph(
                f"<font size=8 color='#{WHITE}'>NET PAY</font><br/>"
                f"<font size=22 color='#{WHITE}'><b>₦{self.payroll.net_pay:,.2f}</b></font>",
                self.styles['Normal']
            ),
            Paragraph(
                f"<font size=8 color='#{SLATE_500}'>Breakdown</font><br/>"
                f"<font size=9 color='#{SLATE_300}'>"
                f"₦{self.payroll.gross_pay:,.2f} gross &ndash; "
                f"₦{self.payroll.total_deductions:,.2f} deductions</font>",
                ParagraphStyle('nptmp', parent=self.styles['Normal'], alignment=TA_RIGHT)
            ),
        ]]

        np_table = Table(net_pay_data, colWidths=[pw * 0.55, pw * 0.45])
        np_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), rc(NAVY)),
            ('ROUNDEDCORNERS', [6, 6, 6, 6]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 18),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 20),
            ('LEFTPADDING', (0, 0), (0, -1), 20),
            ('RIGHTPADDING', (-1, 0), (-1, -1), 20),
        ]))
        story.append(np_table)

        # ── Footer ──────────────────────────────────────────────────
        story.append(Spacer(1, 0.4 * inch))
        story.append(HRFlowable(
            width='100%', thickness=0.25, color=rc(SLATE_300),
            spaceAfter=8, spaceBefore=0
        ))
        story.append(Paragraph(
            "This is a computer-generated payslip and does not require a signature.",
            self.styles['FooterText']
        ))
        story.append(Paragraph(
            f"{self.payroll.client.name} &bull; Confidential",
            self.styles['FooterText']
        ))

        doc.build(story)
        output.seek(0)
        return output
